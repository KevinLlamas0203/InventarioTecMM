# createReporte.py
from flask import Blueprint, request, jsonify, send_file
import psycopg2
import os
import io
import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ReportLab
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, HRFlowable
from reportlab.lib.enums import TA_LEFT

create_reporte_bp = Blueprint("create_reporte_bp", __name__)



def get_connection():
    return psycopg2.connect(os.getenv("DATABASE_URL"))

# ── Helpers de datos ──────────────────────────────────────────────────────────

def get_data_general(conn, rango):
    cur = conn.cursor()
    filtro_sql = {
        "30dias": "WHERE a.fecha_alta >= CURRENT_DATE - INTERVAL '30 days'",
        "3meses": "WHERE a.fecha_alta >= CURRENT_DATE - INTERVAL '3 months'",
        "anio":   "WHERE a.fecha_alta >= DATE_TRUNC('year', CURRENT_DATE)",
        "todo":   ""
    }.get(rango, "")
    cur.execute(f"""
        SELECT a.nombre,
               c.nombre AS categoria,
               e.nombre AS estado,
               u.nombre AS ubicacion,
               COALESCE((us.nombre || ' ' || us.apellido_paterno
                   || COALESCE(' ' || us.apellido_materno, '')), '—') AS asignado_a,
               a.fecha_alta
        FROM activos a
        JOIN categorias  c  ON a.fk_id_categoria = c.id_categoria
        JOIN estados     e  ON a.fk_id_estado    = e.id_estado
        JOIN ubicaciones u  ON a.fk_id_ubicacion = u.id_ubicacion
        LEFT JOIN usuarios us ON a.fk_id_usuario = us.id_usuario
        {filtro_sql}
        ORDER BY a.fecha_alta DESC
    """)
    rows = cur.fetchall()
    cur.close()
    return rows


def get_data_laboratorio(conn, lab):
    cur = conn.cursor()
    if lab == "all":
        cur.execute("""
            SELECT a.nombre,
                   c.nombre AS categoria,
                   e.nombre AS estado,
                   u.nombre AS ubicacion,
                   COALESCE((us.nombre || ' ' || us.apellido_paterno
                       || COALESCE(' ' || us.apellido_materno, '')), '—') AS asignado_a,
                   a.fecha_alta
            FROM activos a
            JOIN categorias  c  ON a.fk_id_categoria = c.id_categoria
            JOIN estados     e  ON a.fk_id_estado    = e.id_estado
            JOIN ubicaciones u  ON a.fk_id_ubicacion = u.id_ubicacion
            LEFT JOIN usuarios us ON a.fk_id_usuario = us.id_usuario
            ORDER BY u.nombre, a.nombre
        """)
    else:
        cur.execute("""
            SELECT a.nombre,
                   c.nombre AS categoria,
                   e.nombre AS estado,
                   u.nombre AS ubicacion,
                   COALESCE((us.nombre || ' ' || us.apellido_paterno
                       || COALESCE(' ' || us.apellido_materno, '')), '—') AS asignado_a,
                   a.fecha_alta
            FROM activos a
            JOIN categorias  c  ON a.fk_id_categoria = c.id_categoria
            JOIN estados     e  ON a.fk_id_estado    = e.id_estado
            JOIN ubicaciones u  ON a.fk_id_ubicacion = u.id_ubicacion
            LEFT JOIN usuarios us ON a.fk_id_usuario = us.id_usuario
            WHERE LOWER(u.nombre) ILIKE %s
            ORDER BY a.nombre
        """, (f"%{lab}%",))
    rows = cur.fetchall()
    cur.close()
    return rows


def get_data_alertas(conn, umbral):
    cur = conn.cursor()
    limite = {"critico": 5, "bajo": 10, "todos": 99999}.get(umbral, 5)
    cur.execute("""
        SELECT con.nombre,
               con.categoria,
               con.stock_actual,
               con.stock_minimo,
               con.ubicacion
        FROM consumibles con
        WHERE con.stock_actual <= %s
        ORDER BY con.stock_actual ASC
    """, (limite,))
    rows = cur.fetchall()
    cur.close()
    return rows

# ── Generador PDF ─────────────────────────────────────────────────────────────

def build_pdf(titulo, subtitulo, columnas, filas, generado_por):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=0.75*inch, rightMargin=0.75*inch,
        topMargin=0.75*inch,  bottomMargin=0.75*inch,
    )
    styles    = getSampleStyleSheet()
    COLOR_HEADER = colors.HexColor("#3f4bf5")
    COLOR_ALT    = colors.HexColor("#f5f5ff")
    COLOR_BORDER = colors.HexColor("#e5e7eb")

    style_title = ParagraphStyle("titulo", parent=styles["Heading1"],
        fontSize=18, textColor=colors.HexColor("#1a1a2e"), spaceAfter=4, alignment=TA_LEFT)
    style_sub  = ParagraphStyle("subtitulo", parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#6b7280"), spaceAfter=2)
    style_meta = ParagraphStyle("meta", parent=styles["Normal"],
        fontSize=9,  textColor=colors.HexColor("#9ca3af"))

    story = [
        Paragraph("Sistema de Inventario ISC", style_sub),
        Paragraph(titulo, style_title),
        Paragraph(subtitulo, style_sub),
        Paragraph(f"Generado por: {generado_por}  ·  Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", style_meta),
        Spacer(1, 0.15*inch),
        HRFlowable(width="100%", thickness=1, color=COLOR_BORDER),
        Spacer(1, 0.2*inch),
        Paragraph(f"Total de registros: <b>{len(filas)}</b>", styles["Normal"]),
        Spacer(1, 0.15*inch),
    ]

    if filas:
        data       = [columnas] + [[str(c) if c is not None else "—" for c in row] for row in filas]
        page_width = letter[0] - 1.5*inch
        col_widths = [page_width / len(columnas)] * len(columnas)
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), COLOR_HEADER),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 9),
            ("ALIGN",         (0,0), (-1,0), "CENTER"),
            ("TOPPADDING",    (0,0), (-1,0), 8),
            ("BOTTOMPADDING", (0,0), (-1,0), 8),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 8),
            ("ALIGN",         (0,1), (-1,-1), "LEFT"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",    (0,1), (-1,-1), 6),
            ("BOTTOMPADDING", (0,1), (-1,-1), 6),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, COLOR_ALT]),
            ("LINEBELOW",     (0,0), (-1,0),  0,   COLOR_HEADER),
            ("LINEBELOW",     (0,1), (-1,-1), 0.3, COLOR_BORDER),
            ("BOX",           (0,0), (-1,-1), 0.5, COLOR_BORDER),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No se encontraron registros para los filtros seleccionados.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer

# ── Generador Excel ───────────────────────────────────────────────────────────

def build_excel(titulo, columnas, filas):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:30]

    header_fill = PatternFill("solid", fgColor="3F4BF5")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill    = PatternFill("solid", fgColor="F5F5FF")

    # Encabezados
    for col_idx, col_name in enumerate(columnas, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = Alignment(horizontal="center", vertical="center")

    # Filas de datos
    for row_idx, row in enumerate(filas, start=2):
        for col_idx, valor in enumerate(row, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx,
                                     value=str(valor) if valor is not None else "—")
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Fila de resumen al final
    ws.append([])
    ws.append([f"Total de registros: {len(filas)}"])
    ws.append([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ── Generador CSV ─────────────────────────────────────────────────────────────

def build_csv(columnas, filas):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columnas)
    for row in filas:
        writer.writerow([str(v) if v is not None else "" for v in row])
    buffer.seek(0)
    # utf-8-sig para que Excel en Windows lo abra correctamente
    return io.BytesIO(buffer.getvalue().encode("utf-8-sig"))

# ── Endpoint principal ────────────────────────────────────────────────────────

@create_reporte_bp.route("/reportes", methods=["POST"])
def create_reporte():
    data = request.get_json()

    titulo       = data.get("titulo",       "Reporte")
    tipo         = data.get("tipo",         "general")
    generado_por = data.get("generado_por", "Sistema")
    fecha        = data.get("fecha",        datetime.now().strftime("%Y-%m-%d"))
    formato      = data.get("formato",      "pdf").lower()

    # Filtros específicos por tipo
    rango  = data.get("rango",  "todo")
    lab    = data.get("lab",    "all")
    umbral = data.get("umbral", "critico")

    # ── Guardar metadato en BD ────────────────────────────────────────────────
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO reportes (titulo, tipo, contenido, generado_por, fecha)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING reporte_id
        """, (titulo, tipo, "", generado_por, fecha))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # ── Obtener datos según tipo ──────────────────────────────────────────────
    try:
        if tipo == "general":
            filas     = get_data_general(conn, rango)
            columnas  = ["Nombre", "Categoría", "Estado", "Ubicación", "Asignado a", "Fecha alta"]
            subtitulo = f"Inventario general · Período: {rango}"

        elif tipo == "laboratorio":
            filas     = get_data_laboratorio(conn, lab)
            columnas  = ["Nombre", "Categoría", "Estado", "Ubicación", "Asignado a", "Fecha alta"]
            subtitulo = f"Laboratorio / área: {lab}"

        elif tipo == "alertas":
            filas     = get_data_alertas(conn, umbral)
            columnas  = ["Nombre", "Categoría", "Stock actual", "Stock mínimo", "Ubicación"]
            subtitulo = f"Umbral: {umbral}"

        else:
            filas, columnas, subtitulo = [], [], ""

        cur.close()
        conn.close()

        # ── Elegir formato y construir archivo ────────────────────────────────
        if formato == "pdf":
            archivo  = build_pdf(titulo, subtitulo, columnas, filas, generado_por)
            mimetype = "application/pdf"
            ext      = "pdf"

        elif formato in ("excel", "xlsx"):
            archivo  = build_excel(titulo, columnas, filas)
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext      = "xlsx"

        elif formato == "csv":
            archivo  = build_csv(columnas, filas)
            mimetype = "text/csv"
            ext      = "csv"

        else:
            return jsonify({"error": f"Formato '{formato}' no soportado"}), 400

        filename = f"reporte_{tipo}_{nuevo_id}_{datetime.now().strftime('%Y%m%d')}.{ext}"
        return send_file(
            archivo,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500